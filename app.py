# app.py
import re
import json
import base64
import html
import datetime as dt

import pandas as pd
import streamlit as st
import openpyxl
import bcrypt

from pathlib import Path

APP_DIR = Path(__file__).parent


TOEGESTAAN_XLSX_PATH = APP_DIR / "toegestaan_gebruik.xlsx"

@st.cache_data(show_spinner=False)
def load_users_df() -> pd.DataFrame:
    if not TOEGESTAAN_XLSX_PATH.exists():
        raise FileNotFoundError(f"Bestand niet gevonden: {TOEGESTAAN_XLSX_PATH.name}")

    df = pd.read_excel(TOEGESTAAN_XLSX_PATH, dtype=str).fillna("")
    df.columns = [c.strip().lower() for c in df.columns]

    # Verwacht: naam, rol, paswoord_hash (aanrader)
    if "naam" not in df.columns or "rol" not in df.columns:
        raise ValueError("Kolommen 'naam' en 'rol' zijn verplicht in toegestaan_gebruik.xlsx")

    # Ondersteun beide: paswoord_hash (aanrader) of paswoord (fallback)
    if "paswoord_hash" not in df.columns and "paswoord" not in df.columns:
        raise ValueError("Voor login heb je 'paswoord_hash' (aanrader) of 'paswoord' nodig.")

    df["naam"] = df["naam"].astype(str).str.strip()
    df["rol"] = df["rol"].astype(str).str.strip().str.lower()

    if "paswoord_hash" in df.columns:
        df["paswoord_hash"] = df["paswoord_hash"].astype(str).str.strip()
    if "paswoord" in df.columns:
        df["paswoord"] = df["paswoord"].astype(str).str.strip()

    # Uniek per naam
    df = df[df["naam"] != ""].copy()
    df = df.drop_duplicates(subset=["naam"], keep="last")

    return df


def verify_password(entered: str, row: pd.Series) -> bool:
    entered = (entered or "").strip()
    if not entered:
        return False

    # Aanrader: bcrypt hash check
    if "paswoord_hash" in row and str(row["paswoord_hash"]).strip():
        try:
            return bcrypt.checkpw(entered.encode("utf-8"), row["paswoord_hash"].encode("utf-8"))
        except Exception:
            return False

    # Fallback (niet ideaal): plain text vergelijken
    if "paswoord" in row and str(row["paswoord"]).strip():
        return entered == str(row["paswoord"]).strip()

    return False


def require_login() -> None:
    if st.session_state.get("auth_ok"):
        return

    st.title("🔐 Inloggen")
    st.caption("Toegang is beveiligd. Meld aan om verder te gaan.")

    users = load_users_df()

    naam = st.text_input("Naam", placeholder="bv. janssens", key="login_naam")
    pw = st.text_input("Paswoord", type="password", key="login_pw")

    c1, c2 = st.columns([1, 2])
    with c1:
        do_login = st.button("Inloggen", use_container_width=True)

    if do_login:
        naam_clean = (naam or "").strip()
        match = users[users["naam"] == naam_clean]

        if match.empty:
            st.error("Onbekende gebruiker.")
            st.stop()

        row = match.iloc[0]
        if verify_password(pw, row):
            st.session_state["auth_ok"] = True
            st.session_state["user_naam"] = row["naam"]
            st.session_state["user_rol"] = row.get("rol", "viewer")
            st.success("Ingelogd.")
            st.rerun()
        else:
            st.error("Onjuist paswoord.")
            st.stop()

    st.stop()


def logout_button() -> None:
    with st.sidebar:
        st.markdown("---")
        st.write(f"👤 **{st.session_state.get('user_naam','')}**")
        st.write(f"🔑 Rol: **{st.session_state.get('user_rol','')}**")
        if st.button("Uitloggen"):
            for k in ["auth_ok", "user_naam", "user_rol"]:
                st.session_state.pop(k, None)
            st.rerun()


# ----------------------------
# Paths / Config
# ----------------------------
APP_DIR = Path(__file__).parent

XLSM_PATH = APP_DIR / "schade met macro.xlsm"
SCHADESHEET = "BRON"

GESPREKKEN_XLSX_PATH = APP_DIR / "Overzicht gesprekken (aangepast).xlsx"
GESPREKKEN_SHEET_NAME = "gesprekken per thema"

COACHINGS_XLSX_PATH = APP_DIR / "Coachingslijst.xlsx"
COACHINGS_SHEET_VOLTOOID = "Voltooide coachings"
COACHINGS_SHEET_COACHING = "Coaching"

PERSONEEL_JSON_PATH = APP_DIR / "personeelsficheGB.json"
LOGO_PATH = APP_DIR / "logo.png"

# External CSS
CSS_PATH = APP_DIR / "styles.css"

# BRON columns to load (including teamcoach from BRON)
SCHADE_COLS = [
    "personeelsnr",
    "volledige naam",
    "teamcoach",
    "Datum",
    "Link",
    "Locatie",
    "voertuig",
    "bus/tram",
    "type",
]

PAGES = [
    ("dashboard", "Dashboard"),
    ("chauffeur", "Chauffeur"),
    ("voertuig", "Voertuig"),
    ("locatie", "Locatie"),
    ("coaching", "Coaching"),
    ("analyse", "Analyse"),
]


# ----------------------------
# Helpers
# ----------------------------
def load_css(path: Path) -> None:
    """Load CSS from external file and inject into Streamlit."""
    if not path.exists():
        st.warning(f"CSS-bestand niet gevonden: {path.name} (zet dit naast app.py)")
        return
    css = path.read_text(encoding="utf-8")
    st.markdown(f"<style>{css}</style>", unsafe_allow_html=True)


def norm(s) -> str:
    return str(s).strip().lower()


def clean_id(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s.strip()


def clean_text(v) -> str:
    return "" if v is None else str(v).strip()


def parse_year(v) -> int | None:
    if v is None:
        return None

    # Excel serial date (soms als getal)
    if isinstance(v, (int, float)) and 30000 < float(v) < 60000:
        try:
            base = dt.datetime(1899, 12, 30)
            d = base + dt.timedelta(days=float(v))
            return d.year
        except Exception:
            pass

    if isinstance(v, (dt.date, dt.datetime)):
        return v.year

    s = str(v).strip()
    if not s:
        return None

    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        return int(m.group(3))

    m2 = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if m2:
        return int(m2.group(1))

    try:
        return dt.datetime.fromisoformat(s).year
    except Exception:
        return None


def format_ddmmyyyy(v) -> str:
    """Toon altijd dd-mm-jjjj; tijd/uurnotatie verdwijnt."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    try:
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            return s
        return ts.strftime("%d-%m-%Y")
    except Exception:
        return s


def img_to_data_uri(path: Path) -> str:
    b = path.read_bytes()
    ext = path.suffix.lower().lstrip(".")
    mime = "png" if ext == "png" else ext
    return f"data:image/{mime};base64,{base64.b64encode(b).decode('utf-8')}"


def _find_col(df: pd.DataFrame, wanted: str) -> str | None:
    w = norm(wanted)

    for c in df.columns:
        if norm(c) == w:
            return c

    if w in ["nummer", "personeelsnr", "personeelsnummer", "p-nr", "p_nr", "p nr", "p-nr."]:
        for alt in [
            "nr",
            "id",
            "persnr",
            "personeelsnr",
            "personeelsnummer",
            "nummer",
            "employeeid",
            "employee_id",
            "p-nr",
            "p nr",
            "p_nr",
            "p-nr.",
            "p-nr (p-nr)",
        ]:
            for c in df.columns:
                if norm(c) == norm(alt):
                    return c

    if w == "datum":
        for alt in ["date", "datum gesprek", "gespreksdatum", "datum coaching", "coachingsdatum"]:
            for c in df.columns:
                if norm(c) == norm(alt):
                    return c

    if w == "info":
        for alt in [
            "informatie",
            "opmerking",
            "opmerkingen",
            "beschrijving",
            "details",
            "thema",
            "onderwerp",
            "samenvatting",
            "actiepunten",
            "resultaat",
            "notities",
            "commentaar",
            "opmerkingen (coach)",
            "opmerkingen chauffeur",
            "opmerkingen",
        ]:
            for c in df.columns:
                if norm(c) == norm(alt):
                    return c

    if w in ["volledige naam", "chauffeurnaam", "naam"]:
        for alt in [
            "chauffeurnaam",
            "chauffeur naam",
            "naam",
            "medewerker",
            "werknemer",
            "chauffeur",
            "volledige naam",
            "full name",
            "fullname",
            "displayname",
            "display_name",
        ]:
            for c in df.columns:
                if norm(c) == norm(alt):
                    return c

    return None


def _flatten_json_to_records(data):
    if data is None:
        return []
    if isinstance(data, list):
        return [x for x in data if isinstance(x, dict)]
    if isinstance(data, dict):
        for k in ["data", "items", "results", "records"]:
            if k in data:
                return _flatten_json_to_records(data[k])
        if data and all(isinstance(v, dict) for v in data.values()):
            out = []
            for key, val in data.items():
                rec = dict(val)
                rec["_key"] = str(key)
                out.append(rec)
            return out
        return [data]
    return []


def render_html_table(
    df: pd.DataFrame,
    col_order: list[str],
    col_widths: dict[str, str],
    max_height_px: int = 520,
) -> None:
    view = df[col_order].copy()
    for c in col_order:
        view[c] = view[c].fillna("").astype(str)

    ths = []
    for c in col_order:
        w = col_widths.get(c, "auto")
        ths.append(f'<th style="width:{w}">{html.escape(c)}</th>')
    thead = "<tr>" + "".join(ths) + "</tr>"

    trs = []
    for _, row in view.iterrows():
        tds = []
        for c in col_order:
            cell = row[c]
            safe = html.escape(cell).replace("\n", "<br/>")
            tds.append(f"<td>{safe}</td>")
        trs.append("<tr>" + "".join(tds) + "</tr>")
    tbody = "".join(trs)

    st.markdown(
        f"""
        <div class="ot-table-wrap" style="max-height:{max_height_px}px;">
          <table class="ot-table">
            <thead>{thead}</thead>
            <tbody>{tbody}</tbody>
          </table>
        </div>
        """,
        unsafe_allow_html=True,
    )


# ----------------------------
# Navigation state
# ----------------------------
def get_page(default="dashboard") -> str:
    try:
        v = st.query_params.get("page", default)
        if isinstance(v, list):
            v = v[0] if v else default
        v = str(v).strip().lower()
    except Exception:
        v = default
    valid = {pid for pid, _ in PAGES}
    return v if v in valid else default


def set_page(page_id: str) -> None:
    st.query_params["page"] = page_id
    st.rerun()


# ----------------------------
# Loaders
# ----------------------------
@st.cache_data(show_spinner=False)
def load_schade_df() -> pd.DataFrame:
    if not XLSM_PATH.exists():
        raise FileNotFoundError(f"Bestand niet gevonden: {XLSM_PATH.name} (zet dit naast app.py)")

    wb = openpyxl.load_workbook(XLSM_PATH, data_only=True, keep_vba=True)
    if SCHADESHEET not in wb.sheetnames:
        raise ValueError(f"Tabblad '{SCHADESHEET}' niet gevonden in {XLSM_PATH.name}")

    ws = wb[SCHADESHEET]

    header = [c.value for c in ws[1]]
    header_map = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        key = norm(h)
        if key and key not in header_map:
            header_map[key] = idx

    def find_idx(col: str) -> int | None:
        key = norm(col)
        if key in header_map:
            return header_map[key]

        if col == "bus/tram":
            for alt in ["bus/ tram", "bus / tram", "bus - tram"]:
                if alt in header_map:
                    return header_map[alt]
        if col == "volledige naam":
            for alt in ["naam", "volledige naam.", "volledige naam "]:
                if alt in header_map:
                    return header_map[alt]
        if col == "teamcoach":
            for alt in ["team coach", "team_coach", "coach", "teamcoach "]:
                if alt in header_map:
                    return header_map[alt]
        return None

    idx_map = {c: find_idx(c) for c in SCHADE_COLS}

    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        obj = {}
        any_val = False

        for col in SCHADE_COLS:
            i = idx_map.get(col)
            val = r[i] if (i is not None and i < len(r)) else None

            if val is not None and str(val).strip() != "":
                any_val = True

            if col == "Datum" and isinstance(val, (dt.date, dt.datetime)):
                val = val.isoformat()

            obj[col] = val

        if any_val:
            rows.append(obj)

    df = pd.DataFrame(rows)
    for c in SCHADE_COLS:
        if c not in df.columns:
            df[c] = None

    df["personeelsnr"] = df["personeelsnr"].apply(clean_id)
    df["volledige naam"] = df["volledige naam"].apply(clean_text)
    df["teamcoach"] = df["teamcoach"].apply(clean_text)
    df["voertuig"] = df["voertuig"].apply(clean_text)

    df["_jaar"] = df["Datum"].apply(parse_year)
    df["_search"] = (
        df["personeelsnr"].fillna("").astype(str)
        + " "
        + df["volledige naam"].fillna("").astype(str)
        + " "
        + df["teamcoach"].fillna("").astype(str)
        + " "
        + df["voertuig"].fillna("").astype(str)
    ).str.lower()

    return df


@st.cache_data(show_spinner=False)
def load_gesprekken_df() -> pd.DataFrame:
    if not GESPREKKEN_XLSX_PATH.exists():
        raise FileNotFoundError(f"Bestand niet gevonden: {GESPREKKEN_XLSX_PATH.name} (zet dit naast app.py)")

    xls = pd.ExcelFile(GESPREKKEN_XLSX_PATH)
    if GESPREKKEN_SHEET_NAME not in xls.sheet_names:
        raise ValueError(
            f"Tabblad '{GESPREKKEN_SHEET_NAME}' niet gevonden in {GESPREKKEN_XLSX_PATH.name}. "
            f"Gevonden tabs: {xls.sheet_names}"
        )

    df = pd.read_excel(GESPREKKEN_XLSX_PATH, sheet_name=GESPREKKEN_SHEET_NAME, dtype=str)

    num_col = _find_col(df, "nummer")
    date_col = _find_col(df, "Datum")
    info_col = _find_col(df, "Info")
    name_col = _find_col(df, "Chauffeurnaam")

    if num_col is None:
        raise ValueError("Kolom 'nummer' (personeelsnr) niet gevonden in 'gesprekken per thema'.")

    if num_col != "nummer":
        df = df.rename(columns={num_col: "nummer"})
    if date_col and date_col != "Datum":
        df = df.rename(columns={date_col: "Datum"})
    if info_col and info_col != "Info":
        df = df.rename(columns={info_col: "Info"})
    if name_col and name_col != "Chauffeurnaam":
        df = df.rename(columns={name_col: "Chauffeurnaam"})

    for c in ["Datum", "Info", "Chauffeurnaam"]:
        if c not in df.columns:
            df[c] = ""

    df["nummer"] = df["nummer"].apply(clean_id)
    df["Datum"] = df["Datum"].apply(clean_text)
    df["Info"] = df["Info"].apply(clean_text)
    df["Chauffeurnaam"] = df["Chauffeurnaam"].apply(clean_text)

    df["_search"] = (
        df["nummer"].fillna("").astype(str)
        + " "
        + df["Chauffeurnaam"].fillna("").astype(str)
        + " "
        + df["Info"].fillna("").astype(str)
    ).str.lower()
    df["_jaar"] = df["Datum"].apply(parse_year)
    return df


@st.cache_data(show_spinner=False)
def load_coaching_voltooid_df() -> pd.DataFrame:
    if not COACHINGS_XLSX_PATH.exists():
        return pd.DataFrame(columns=["nummer", "Chauffeurnaam", "Datum", "Info", "_search", "_jaar"])

    xls = pd.ExcelFile(COACHINGS_XLSX_PATH)
    if COACHINGS_SHEET_VOLTOOID not in xls.sheet_names:
        raise ValueError(
            f"Tabblad '{COACHINGS_SHEET_VOLTOOID}' niet gevonden in {COACHINGS_XLSX_PATH.name}. "
            f"Gevonden tabs: {xls.sheet_names}"
        )

    df = pd.read_excel(COACHINGS_XLSX_PATH, sheet_name=COACHINGS_SHEET_VOLTOOID, dtype=str)

    num_col = _find_col(df, "nummer") or _find_col(df, "personeelsnr")
    name_col = _find_col(df, "Chauffeurnaam") or _find_col(df, "naam") or _find_col(df, "volledige naam")
    date_col = _find_col(df, "Datum")
    info_col = _find_col(df, "Info")

    if num_col is None:
        df["nummer"] = ""
    else:
        if num_col != "nummer":
            df = df.rename(columns={num_col: "nummer"})

    if name_col is None:
        df["Chauffeurnaam"] = ""
    else:
        if name_col != "Chauffeurnaam":
            df = df.rename(columns={name_col: "Chauffeurnaam"})

    if date_col is None:
        df["Datum"] = ""
    else:
        if date_col != "Datum":
            df = df.rename(columns={date_col: "Datum"})

    if info_col is None:
        candidates = []
        for c in df.columns:
            if norm(c) in [
                "thema",
                "onderwerp",
                "opmerking",
                "opmerkingen",
                "samenvatting",
                "notities",
                "commentaar",
                "actiepunten",
                "resultaat",
            ]:
                candidates.append(c)
        if candidates:
            df["Info"] = df[candidates].fillna("").astype(str).agg(" | ".join, axis=1)
        else:
            df["Info"] = ""
    else:
        if info_col != "Info":
            df = df.rename(columns={info_col: "Info"})

    df["nummer"] = df["nummer"].apply(clean_id)
    df["Chauffeurnaam"] = df["Chauffeurnaam"].apply(clean_text)
    df["Datum"] = df["Datum"].apply(clean_text)
    df["Info"] = df["Info"].apply(clean_text)

    df["_search"] = (
        df["nummer"].fillna("").astype(str)
        + " "
        + df["Chauffeurnaam"].fillna("").astype(str)
        + " "
        + df["Info"].fillna("").astype(str)
    ).str.lower()
    df["_jaar"] = df["Datum"].apply(parse_year)
    return df


@st.cache_data(show_spinner=False)
def load_coaching_tab_df() -> pd.DataFrame:
    """
    Coachingslijst.xlsx -> tab 'Coaching'
    Kolommen: P-nr, Volledige naam, Opmerkingen
    """
    if not COACHINGS_XLSX_PATH.exists():
        return pd.DataFrame(columns=["nummer", "Chauffeurnaam", "Info", "_search"])

    xls = pd.ExcelFile(COACHINGS_XLSX_PATH)
    if COACHINGS_SHEET_COACHING not in xls.sheet_names:
        raise ValueError(
            f"Tabblad '{COACHINGS_SHEET_COACHING}' niet gevonden in {COACHINGS_XLSX_PATH.name}. "
            f"Gevonden tabs: {xls.sheet_names}"
        )

    df = pd.read_excel(COACHINGS_XLSX_PATH, sheet_name=COACHINGS_SHEET_COACHING, dtype=str)

    pnr_col = _find_col(df, "P-nr") or _find_col(df, "nummer") or _find_col(df, "personeelsnr")
    name_col = _find_col(df, "Volledige naam") or _find_col(df, "naam") or _find_col(df, "chauffeurnaam")
    opm_col = _find_col(df, "Opmerkingen") or _find_col(df, "Info")

    if pnr_col is None:
        df["nummer"] = ""
    else:
        if pnr_col != "nummer":
            df = df.rename(columns={pnr_col: "nummer"})

    if name_col is None:
        df["Chauffeurnaam"] = ""
    else:
        if name_col != "Chauffeurnaam":
            df = df.rename(columns={name_col: "Chauffeurnaam"})

    if opm_col is None:
        df["Info"] = ""
    else:
        if opm_col != "Info":
            df = df.rename(columns={opm_col: "Info"})

    df["nummer"] = df["nummer"].apply(clean_id)
    df["Chauffeurnaam"] = df["Chauffeurnaam"].apply(clean_text)
    df["Info"] = df["Info"].apply(clean_text)

    df["_search"] = (
        df["nummer"].fillna("").astype(str)
        + " "
        + df["Chauffeurnaam"].fillna("").astype(str)
        + " "
        + df["Info"].fillna("").astype(str)
    ).str.lower()
    return df


@st.cache_data(show_spinner=False)
def load_personeelsfiche_df() -> pd.DataFrame:
    if not PERSONEEL_JSON_PATH.exists():
        return pd.DataFrame(columns=["_search"])

    try:
        data = json.loads(PERSONEEL_JSON_PATH.read_text(encoding="utf-8"))
    except Exception:
        data = json.loads(PERSONEEL_JSON_PATH.read_text())

    records = _flatten_json_to_records(data)
    if not records:
        return pd.DataFrame(columns=["_search"])

    df = pd.DataFrame(records)

    id_col = _find_col(df, "personeelsnr") or _find_col(df, "nummer") or _find_col(df, "personeelsnummer")
    name_col = _find_col(df, "volledige naam") or _find_col(df, "naam") or _find_col(df, "chauffeurnaam")

    if id_col is None and "_key" in df.columns:
        id_col = "_key"

    if id_col and id_col != "personeelsnr":
        df = df.rename(columns={id_col: "personeelsnr"})
        id_col = "personeelsnr"
    if name_col and name_col != "naam":
        df = df.rename(columns={name_col: "naam"})
        name_col = "naam"

    if id_col is None:
        df["personeelsnr"] = ""
        id_col = "personeelsnr"
    if name_col is None:
        df["naam"] = ""
        name_col = "naam"

    df[id_col] = df[id_col].apply(clean_id)
    df[name_col] = df[name_col].apply(clean_text)

    extra_cols = []
    for c in df.columns:
        if c in ["_search", id_col, name_col]:
            continue
        if norm(c) in ["dienst", "afdeling", "team", "functie", "rol", "standplaats", "locatie", "teamcoach"]:
            extra_cols.append(c)

    parts = [df[id_col].fillna("").astype(str), df[name_col].fillna("").astype(str)]
    for c in extra_cols[:6]:
        parts.append(df[c].fillna("").astype(str))

    df["_search"] = parts[0]
    for s in parts[1:]:
        df["_search"] = df["_search"].astype(str) + " " + s.astype(str)
    df["_search"] = df["_search"].str.lower()
    return df


# ----------------------------
# Streamlit setup
# ----------------------------
st.set_page_config(page_title="Analyse en rapportering OT Gent", layout="wide")
load_css(CSS_PATH)

st.set_page_config(page_title="Analyse en rapportering OT Gent", layout="wide")
load_css(CSS_PATH)

require_login()
logout_button()

# daarna pas:
# df_schade = load_schade_df()
# ...


# ----------------------------
# Load data
# ----------------------------
try:
    df_schade = load_schade_df()
except Exception as e:
    st.error(f"Kan schade-data niet laden: {e}")
    st.stop()

try:
    df_gesprekken = load_gesprekken_df()
except Exception as e:
    st.warning(f"Gesprekkenbestand niet geladen: {e}")
    df_gesprekken = pd.DataFrame(columns=["nummer", "Chauffeurnaam", "Datum", "Info", "_search", "_jaar"])

try:
    df_coach_voltooid = load_coaching_voltooid_df()
except Exception as e:
    st.warning(f"Voltooide coachings niet geladen: {e}")
    df_coach_voltooid = pd.DataFrame(columns=["nummer", "Chauffeurnaam", "Datum", "Info", "_search", "_jaar"])

try:
    df_coach_tab = load_coaching_tab_df()
except Exception as e:
    st.warning(f"Tabblad 'Coaching' niet geladen: {e}")
    df_coach_tab = pd.DataFrame(columns=["nummer", "Chauffeurnaam", "Info", "_search"])

df_personeel = load_personeelsfiche_df()

years_schade = df_schade["_jaar"].dropna().unique().tolist() if "_jaar" in df_schade.columns else []
years_gespr = df_gesprekken["_jaar"].dropna().unique().tolist() if "_jaar" in df_gesprekken.columns else []
years_volt = df_coach_voltooid["_jaar"].dropna().unique().tolist() if "_jaar" in df_coach_voltooid.columns else []
years = sorted({int(y) for y in (years_schade + years_gespr + years_volt) if y is not None}, reverse=True)

current_page = get_page("dashboard")


# ----------------------------
# Topbar
# ----------------------------
st.markdown('<div class="ot-topbar">', unsafe_allow_html=True)

c1, c2, c3 = st.columns([2.3, 1.2, 3.5], vertical_alignment="center")

with c1:
    logo_html = f'<img class="ot-logo" src="{img_to_data_uri(LOGO_PATH)}" alt="Logo" />' if LOGO_PATH.exists() else ""
    st.markdown(
        f"""
        <div class="ot-brand">
          {logo_html}
          <div>
            <div class="ot-title">Analyse en rapportering OT Gent</div>
            <div class="ot-sub">schade</div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

with c2:
    year_choice = st.selectbox("Jaar", ["Alle"] + [str(y) for y in years], index=0)

with c3:
    tab_cols = st.columns([1, 1, 1, 1, 1, 1, 0.95], gap="small")

    for (pid, label), col in zip(PAGES, tab_cols[:6]):
        with col:
            active = (pid == current_page)
            st.markdown(f'<div class="ot-tab-btn {"active" if active else ""}">', unsafe_allow_html=True)
            if st.button(label, key=f"tab_{pid}", use_container_width=True):
                set_page(pid)
            st.markdown("</div>", unsafe_allow_html=True)

    with tab_cols[6]:
        st.markdown('<div class="ot-tab-btn">', unsafe_allow_html=True)
        if st.button("↻ Herladen", key="reload_btn", use_container_width=True):
            st.cache_data.clear()
            try:
                st.cache_resource.clear()
            except Exception:
                pass
            st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

st.markdown("</div>", unsafe_allow_html=True)


# ----------------------------
# Year filter views
# ----------------------------
df_schade_view = df_schade[df_schade["_jaar"] == int(year_choice)].copy() if year_choice != "Alle" else df_schade.copy()
df_gesprekken_view = df_gesprekken[df_gesprekken["_jaar"] == int(year_choice)].copy() if year_choice != "Alle" else df_gesprekken.copy()
df_coach_voltooid_view = (
    df_coach_voltooid[df_coach_voltooid["_jaar"] == int(year_choice)].copy()
    if year_choice != "Alle"
    else df_coach_voltooid.copy()
)
# df_coach_tab heeft geen jaarfilter (geen datumkolom)


# ----------------------------
# Pages
# ----------------------------
if current_page == "dashboard":
    st.subheader("Dashboard")

    q = st.text_input(
        "Zoek op personeelsnr of naam.",
        placeholder="Typ om te zoeken…",
    ).strip().lower()

    if not q:
        st.caption("Typ iets in het zoekveld om resultaten te zien.")
        st.stop()

    schade_hits = df_schade_view[df_schade_view["_search"].str.contains(re.escape(q), na=False)].copy()
    gesprekken_hits = df_gesprekken_view[df_gesprekken_view["_search"].str.contains(re.escape(q), na=False)].copy()
    coach_volt_hits = df_coach_voltooid_view[df_coach_voltooid_view["_search"].str.contains(re.escape(q), na=False)].copy()

    coach_tab_hits = pd.DataFrame()
    if "_search" in df_coach_tab.columns and len(df_coach_tab) > 0:
        coach_tab_hits = df_coach_tab[df_coach_tab["_search"].str.contains(re.escape(q), na=False)].copy()

    personeels_hits = pd.DataFrame()
    if "_search" in df_personeel.columns and len(df_personeel) > 0:
        personeels_hits = df_personeel[df_personeel["_search"].str.contains(re.escape(q), na=False)].copy()

    # Personeelsfiche
    st.markdown("#### Personeelsfiche")
    if len(personeels_hits) == 0:
        st.caption("Geen personeelsfiche gevonden voor deze zoekterm.")
    else:
        summary_cols = [c for c in ["personeelsnr", "naam"] if c in personeels_hits.columns]
        if summary_cols:
            st.dataframe(personeels_hits[summary_cols].head(20), use_container_width=True, hide_index=True)

        max_show = 10
        for i, (_, row) in enumerate(personeels_hits.head(max_show).iterrows(), start=1):
            pid = row.get("personeelsnr", "")
            nm = row.get("naam", "")
            title = f"{i}. {pid} — {nm}".strip(" —")
            with st.expander(title, expanded=(i == 1)):
                rec = row.drop(labels=["_search"], errors="ignore").to_dict()
                st.json(rec)

        if len(personeels_hits) > max_show:
            st.caption(f"… en nog {len(personeels_hits) - max_show} extra matches.")

    # Schade
    st.markdown("#### Schade (BRON)")
    if len(schade_hits) == 0:
        st.caption("Geen schadegevallen gevonden voor deze zoekterm.")
    else:
        show_cols = [c for c in SCHADE_COLS if c in schade_hits.columns]
        show = schade_hits[show_cols].head(500).copy()
        if "Datum" in show.columns:
            show["Datum"] = show["Datum"].apply(format_ddmmyyyy)

        column_config = {
            "personeelsnr": st.column_config.TextColumn("personeelsnr", width="small"),
            "volledige naam": st.column_config.TextColumn("volledige naam", width="medium"),
            "teamcoach": st.column_config.TextColumn("teamcoach", width="medium"),
            "Datum": st.column_config.TextColumn("Datum", width="small"),
            "Link": st.column_config.LinkColumn("Open EAF", display_text="Open EAF", width="small"),
            "Locatie": st.column_config.TextColumn("Locatie", width="medium"),
            "voertuig": st.column_config.TextColumn("voertuig", width="medium"),
            "bus/tram": st.column_config.TextColumn("bus/tram", width="small"),
            "type": st.column_config.TextColumn("type", width="small"),
        }

        st.dataframe(
            show,
            use_container_width=True,
            hide_index=True,
            column_config=column_config,
        )

    # Geplande coaching
    st.markdown("#### Geplande coaching")
    if len(coach_tab_hits) == 0:
        st.caption("Geen geplande coaching-info gevonden voor deze zoekterm.")
    else:
        display_ct = coach_tab_hits[["nummer", "Chauffeurnaam", "Info"]].copy()
        render_html_table(
            display_ct.head(300),
            col_order=["nummer", "Chauffeurnaam", "Info"],
            col_widths={"nummer": "90px", "Chauffeurnaam": "220px", "Info": "auto"},
            max_height_px=520,
        )

    # Voltooide coaching
    st.markdown("#### Voltooide coaching")
    if len(coach_volt_hits) == 0:
        st.caption("Geen voltooide coachings gevonden voor deze zoekterm.")
    else:
        display_v = coach_volt_hits[["nummer", "Chauffeurnaam", "Datum", "Info"]].copy()
        display_v["Datum"] = display_v["Datum"].apply(format_ddmmyyyy)
        render_html_table(
            display_v.head(300),
            col_order=["nummer", "Chauffeurnaam", "Datum", "Info"],
            col_widths={"nummer": "90px", "Chauffeurnaam": "180px", "Datum": "120px", "Info": "auto"},
            max_height_px=520,
        )

    # Overzicht gesprekken
    st.markdown("#### Overzicht gesprekken")
    if len(gesprekken_hits) == 0:
        st.caption("Geen gesprekken gevonden voor deze zoekterm.")
    else:
        display_g = gesprekken_hits[["nummer", "Chauffeurnaam", "Datum", "Info"]].copy()
        display_g["Datum"] = display_g["Datum"].apply(format_ddmmyyyy)
        render_html_table(
            display_g.head(300),
            col_order=["nummer", "Chauffeurnaam", "Datum", "Info"],
            col_widths={"nummer": "90px", "Chauffeurnaam": "180px", "Datum": "120px", "Info": "auto"},
            max_height_px=520,
        )

elif current_page == "chauffeur":
    st.subheader("Chauffeur")

    if df_schade_view.empty:
        st.info("Geen schadegegevens beschikbaar voor deze selectie.")
        st.stop()

    # Controls
    top_n = st.selectbox("Top", [10, 20, 50, 100], index=1)
    min_aantal = st.slider("Minimum aantal schadegevallen", 1, 20, 1)

    # ---- Top chauffeurs ----
    st.markdown("### 🚗 Chauffeurs met meeste schadegevallen")

    top_chauffeurs = (
        df_schade_view
        .groupby(["personeelsnr", "volledige naam"], dropna=False)
        .size()
        .reset_index(name="Aantal schadegevallen")
        .sort_values("Aantal schadegevallen", ascending=False)
    )

    top_chauffeurs_filtered = top_chauffeurs[top_chauffeurs["Aantal schadegevallen"] >= min_aantal].head(top_n)

    st.dataframe(
        top_chauffeurs_filtered,
        use_container_width=True,
        hide_index=True,
        column_config={
            "personeelsnr": st.column_config.TextColumn("Personeelsnr", width="small"),
            "volledige naam": st.column_config.TextColumn("Chauffeur", width="medium"),
            "Aantal schadegevallen": st.column_config.NumberColumn("Aantal", width="small"),
        },
    )

    if len(top_chauffeurs_filtered) == 0:
        st.caption("Geen chauffeurs binnen deze filters.")

    # ---- Teamcoach ----
    st.markdown("### 👥 Teamcoach: aantal schadegevallen")

    if "teamcoach" not in df_schade_view.columns:
        st.warning("Kolom 'teamcoach' niet gevonden in BRON.")
        st.stop()

    schade_per_teamcoach = (
        df_schade_view
        .assign(teamcoach=df_schade_view["teamcoach"].fillna("").astype(str).str.strip())
        .replace({"teamcoach": {"": "(onbekend)"}})
        .groupby("teamcoach", dropna=False)
        .size()
        .reset_index(name="Aantal schadegevallen")
        .sort_values("Aantal schadegevallen", ascending=False)
    )

    st.dataframe(
        schade_per_teamcoach.rename(columns={"Aantal schadegevallen": "Aantal"}),
        use_container_width=True,
        hide_index=True,
        column_config={
            "teamcoach": st.column_config.TextColumn("Teamcoach", width="medium"),
            "Aantal": st.column_config.NumberColumn("Aantal", width="small"),
        },
    )

    # Sorted bar chart (hoog -> laag)
    schade_per_teamcoach_sorted = (
        schade_per_teamcoach
        .sort_values("Aantal schadegevallen", ascending=False)
        .set_index("teamcoach")
    )

    st.bar_chart(
        schade_per_teamcoach_sorted["Aantal schadegevallen"]
    )

elif current_page == "voertuig":
    st.subheader("Voertuig")

    if df_schade_view.empty:
        st.info("Geen schadegegevens beschikbaar voor deze selectie.")
        st.stop()

    # ----------------------------
    # Lokale helper: robuuste datum naar maand (werkt met string/iso/datetime)
    # ----------------------------
    def _to_month(v) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        if not s:
            return ""
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            return ""
        return ts.strftime("%Y-%m")

    # ----------------------------
    # Controls
    # ----------------------------
    c1, c2, c3, c4 = st.columns([1.0, 1.0, 1.1, 1.4])
    with c1:
        top_n = st.selectbox("Top", [10, 20, 50, 100, 200], index=1)
    with c2:
        min_aantal = st.slider("Minimum aantal schadegevallen", 1, 50, 1)
    with c3:
        # bus/tram filter (optioneel)
        bt_vals = (
            df_schade_view["bus/tram"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "(onbekend)")
            .unique()
            .tolist()
        )
        bt_vals = sorted(bt_vals)
        bus_tram = st.selectbox("Bus/Tram", ["Alles"] + bt_vals, index=0)
    with c4:
        voertuig_q = st.text_input("Zoek voertuig", placeholder="bv. 6301, 7205, ...").strip().lower()

    tmp = df_schade_view.copy()

    # Normaliseer kernkolommen
    tmp["voertuig"] = tmp["voertuig"].fillna("").astype(str).str.strip()
    tmp["bus/tram"] = tmp["bus/tram"].fillna("").astype(str).str.strip().replace("", "(onbekend)")
    tmp["Locatie"] = tmp["Locatie"].fillna("").astype(str).str.strip()
    tmp["type"] = tmp["type"].fillna("").astype(str).str.strip()
    tmp["teamcoach"] = tmp["teamcoach"].fillna("").astype(str).str.strip()

    # Filter bus/tram
    if bus_tram != "Alles":
        tmp = tmp[tmp["bus/tram"] == bus_tram].copy()

    # Filter voertuig zoekterm
    if voertuig_q:
        tmp = tmp[tmp["voertuig"].str.lower().str.contains(re.escape(voertuig_q), na=False)].copy()

    # Lege voertuigen labelen
    tmp["voertuig"] = tmp["voertuig"].replace("", "(onbekend)")

    # ----------------------------
    # KPI’s
    # ----------------------------
    total_cases = len(tmp)
    unique_voertuigen = tmp["voertuig"].nunique(dropna=True)
    avg_per_voertuig = (total_cases / unique_voertuigen) if unique_voertuigen else 0.0

    # top voertuig
    top_voertuig = ""
    top_voertuig_count = 0
    if total_cases > 0:
        vc = tmp.groupby("voertuig").size().sort_values(ascending=False)
        if len(vc) > 0:
            top_voertuig = str(vc.index[0])
            top_voertuig_count = int(vc.iloc[0])

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Schadegevallen", f"{total_cases}")
    k2.metric("Unieke voertuigen", f"{unique_voertuigen}")
    k3.metric("Gemiddeld / voertuig", f"{avg_per_voertuig:.2f}")
    k4.metric("Top voertuig", f"{top_voertuig_count} — {top_voertuig}" if top_voertuig else "—")

    st.divider()

    # ----------------------------
    # Top voertuigen (tabel + bar chart)
    # ----------------------------
    st.markdown("### 🚋 Top voertuigen met meeste schadegevallen")

    voertuigen_counts = (
        tmp.groupby("voertuig")
        .size()
        .reset_index(name="Aantal")
        .sort_values("Aantal", ascending=False)
    )

    voertuigen_counts = voertuigen_counts[voertuigen_counts["Aantal"] >= min_aantal].copy()

    # Voeg extra kolommen toe: meest voorkomende bus/tram, laatste datum, top locatie
    # (alles optioneel maar handig)
    def _mode_or_empty(s: pd.Series) -> str:
        s = s.dropna().astype(str).str.strip()
        s = s[s != ""]
        if s.empty:
            return ""
        return s.value_counts().index[0]

    extra = (
        tmp.groupby("voertuig", dropna=False)
        .agg(
            BusTram=("bus/tram", _mode_or_empty),
            LaatsteDatum=("Datum", lambda x: pd.to_datetime(x, dayfirst=True, errors="coerce").max()),
            TopLocatie=("Locatie", _mode_or_empty),
        )
        .reset_index()
    )
    
    # Forceer datetime dtype => .dt werkt altijd
    extra["LaatsteDatum"] = pd.to_datetime(extra["LaatsteDatum"], errors="coerce")
    extra["LaatsteDatum"] = extra["LaatsteDatum"].dt.strftime("%d-%m-%Y").fillna("")


    top_table = voertuigen_counts.merge(extra, on="voertuig", how="left").head(top_n)

    st.dataframe(
        top_table,
        use_container_width=True,
        hide_index=True,
        column_config={
            "voertuig": st.column_config.TextColumn("Voertuig", width="medium"),
            "Aantal": st.column_config.NumberColumn("Aantal", width="small"),
            "BusTram": st.column_config.TextColumn("Bus/Tram (meest voork.)", width="small"),
            "LaatsteDatum": st.column_config.TextColumn("Laatste datum", width="small"),
            "TopLocatie": st.column_config.TextColumn("Top locatie", width="medium"),
        },
    )

    if top_table.empty:
        st.caption("Geen voertuigen binnen deze filters.")
        st.stop()

    # Bar chart (hoog -> laag)
    chart_df = top_table.set_index("voertuig")["Aantal"]
    st.bar_chart(chart_df)

    st.divider()

    # ----------------------------
    # Kies voertuig + trend per maand + details
    # ----------------------------
    st.markdown("### 📈 Trend & details voor gekozen voertuig")

    voertuig_options = top_table["voertuig"].tolist()
    default_idx = 0
    gekozen_voertuig = st.selectbox("Kies voertuig", voertuig_options, index=default_idx)

    vdf = tmp[tmp["voertuig"] == gekozen_voertuig].copy()

    # trend per maand
    vdf["Maand"] = vdf["Datum"].apply(_to_month)
    per_maand = (
        vdf[vdf["Maand"] != ""]
        .groupby("Maand")
        .size()
        .reset_index(name="Aantal")
        .sort_values("Maand")
    )

    cL, cR = st.columns([1.2, 1.0], gap="large")
    with cL:
        st.markdown("#### Schade per maand")
        if per_maand.empty:
            st.caption("Geen geldige datums om per maand te groeperen.")
        else:
            st.dataframe(per_maand, use_container_width=True, hide_index=True)
            st.bar_chart(per_maand.set_index("Maand")["Aantal"])

    with cR:
        st.markdown("#### Breakdown (top 10)")
        # type
        per_type = (
            vdf.assign(type=vdf["type"].replace("", "(onbekend)"))
            .groupby("type")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.caption("Type")
        st.dataframe(per_type, use_container_width=True, hide_index=True)

        # locatie
        per_loc = (
            vdf.assign(Locatie=vdf["Locatie"].replace("", "(onbekend)"))
            .groupby("Locatie")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.caption("Locatie")
        st.dataframe(per_loc, use_container_width=True, hide_index=True)

    st.markdown("#### Detail-lijst (laatste 200)")
    detail_cols = [c for c in ["Datum", "Locatie", "type", "bus/tram", "teamcoach", "volledige naam", "personeelsnr", "Link"] if c in vdf.columns]
    details = vdf[detail_cols].copy()

    if "Datum" in details.columns:
        details["Datum"] = details["Datum"].apply(format_ddmmyyyy)

    # Sorteer (nieuwste eerst) op Datum indien mogelijk
    try:
        sort_ts = pd.to_datetime(vdf["Datum"].astype(str), dayfirst=True, errors="coerce")
        details["_sort"] = sort_ts
        details = details.sort_values("_sort", ascending=False).drop(columns=["_sort"])
    except Exception:
        pass

    # LinkColumn netjes houden
    if "Link" in details.columns:
        details["Link"] = details["Link"].replace({"": None})

    column_config = {}
    if "Link" in details.columns:
        column_config["Link"] = st.column_config.LinkColumn("Open EAF", display_text="Open EAF", width="small")

    st.dataframe(
        details.head(200),
        use_container_width=True,
        hide_index=True,
        column_config=column_config if column_config else None,
    )



elif current_page == "locatie":
    st.subheader("Locatie")

    if df_schade_view.empty:
        st.info("Geen schadegegevens beschikbaar voor deze selectie.")
        st.stop()

    # ----------------------------
    # Lokale helper: datum -> maand (YYYY-MM), robuust genoeg voor strings/iso
    # ----------------------------
    def _to_month(v) -> str:
        if v is None:
            return ""
        s = str(v).strip()
        if not s:
            return ""
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        if pd.isna(ts):
            return ""
        return ts.strftime("%Y-%m")

    # ----------------------------
    # Controls
    # ----------------------------
    c1, c2, c3, c4, c5 = st.columns([1.0, 1.1, 1.1, 1.2, 1.4])
    with c1:
        top_n = st.selectbox("Top", [10, 20, 50, 100, 200], index=1)
    with c2:
        min_aantal = st.slider("Minimum aantal", 1, 50, 1)
    with c3:
        bt_vals = (
            df_schade_view["bus/tram"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "(onbekend)")
            .unique()
            .tolist()
        )
        bt_vals = sorted(bt_vals)
        bus_tram = st.selectbox("Bus/Tram", ["Alles"] + bt_vals, index=0)
    with c4:
        type_vals = (
            df_schade_view["type"]
            .fillna("")
            .astype(str)
            .str.strip()
            .replace("", "(onbekend)")
            .unique()
            .tolist()
        )
        type_vals = sorted(type_vals)
        type_filter = st.selectbox("Type", ["Alles"] + type_vals, index=0)
    with c5:
        locatie_q = st.text_input("Zoek locatie", placeholder="bv. Gent, stelplaats, ...").strip().lower()

    tmp = df_schade_view.copy()

    # Normaliseer kernkolommen
    tmp["Locatie"] = tmp["Locatie"].fillna("").astype(str).str.strip()
    tmp["bus/tram"] = tmp["bus/tram"].fillna("").astype(str).str.strip().replace("", "(onbekend)")
    tmp["type"] = tmp["type"].fillna("").astype(str).str.strip().replace("", "(onbekend)")
    tmp["voertuig"] = tmp["voertuig"].fillna("").astype(str).str.strip().replace("", "(onbekend)")
    tmp["teamcoach"] = tmp["teamcoach"].fillna("").astype(str).str.strip().replace("", "(onbekend)")

    # Label lege locaties
    tmp["Locatie"] = tmp["Locatie"].replace("", "(onbekend)")

    # Filters
    if bus_tram != "Alles":
        tmp = tmp[tmp["bus/tram"] == bus_tram].copy()

    if type_filter != "Alles":
        tmp = tmp[tmp["type"] == type_filter].copy()

    if locatie_q:
        tmp = tmp[tmp["Locatie"].str.lower().str.contains(re.escape(locatie_q), na=False)].copy()

    # ----------------------------
    # KPI’s
    # ----------------------------
    total_cases = len(tmp)
    unique_locaties = tmp["Locatie"].nunique(dropna=True)
    avg_per_loc = (total_cases / unique_locaties) if unique_locaties else 0.0

    top_loc = ""
    top_loc_count = 0
    if total_cases > 0:
        lc = tmp.groupby("Locatie").size().sort_values(ascending=False)
        if len(lc) > 0:
            top_loc = str(lc.index[0])
            top_loc_count = int(lc.iloc[0])

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Schadegevallen", f"{total_cases}")
    k2.metric("Unieke locaties", f"{unique_locaties}")
    k3.metric("Gemiddeld / locatie", f"{avg_per_loc:.2f}")
    k4.metric("Top locatie", f"{top_loc_count} — {top_loc}" if top_loc else "—")

    st.divider()

    # ----------------------------
    # Top locaties (tabel + bar chart)
    # ----------------------------
    st.markdown("### 📍 Hotspots: locaties met meeste schadegevallen")

    locaties_counts = (
        tmp.groupby("Locatie")
        .size()
        .reset_index(name="Aantal")
        .sort_values("Aantal", ascending=False)
    )
    locaties_counts = locaties_counts[locaties_counts["Aantal"] >= min_aantal].copy()

    def _mode_or_empty(s: pd.Series) -> str:
        s = s.dropna().astype(str).str.strip()
        s = s[s != ""]
        if s.empty:
            return ""
        return s.value_counts().index[0]

    extra = (
        tmp.groupby("Locatie", dropna=False)
        .agg(
            LaatsteDatum=("Datum", lambda x: pd.to_datetime(x, dayfirst=True, errors="coerce").max()),
            TopType=("type", _mode_or_empty),
            TopVoertuig=("voertuig", _mode_or_empty),
            TopTeamcoach=("teamcoach", _mode_or_empty),
        )
        .reset_index()
    )

    # Forceer datetime dtype => .dt werkt altijd
    extra["LaatsteDatum"] = pd.to_datetime(extra["LaatsteDatum"], errors="coerce")
    extra["LaatsteDatum"] = extra["LaatsteDatum"].dt.strftime("%d-%m-%Y").fillna("")

    top_table = locaties_counts.merge(extra, on="Locatie", how="left").head(top_n)

    st.dataframe(
        top_table,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Locatie": st.column_config.TextColumn("Locatie", width="large"),
            "Aantal": st.column_config.NumberColumn("Aantal", width="small"),
            "LaatsteDatum": st.column_config.TextColumn("Laatste datum", width="small"),
            "TopType": st.column_config.TextColumn("Meest voorkomend type", width="medium"),
            "TopVoertuig": st.column_config.TextColumn("Meest voorkomend voertuig", width="small"),
            "TopTeamcoach": st.column_config.TextColumn("Meest voorkomend teamcoach", width="medium"),
        },
    )

    if top_table.empty:
        st.caption("Geen locaties binnen deze filters.")
        st.stop()

    st.divider()

    # ----------------------------
    # Kies locatie + trend per maand + breakdown + details
    # ----------------------------
    st.markdown("### 📈 Trend & details voor gekozen locatie")

    locatie_options = top_table["Locatie"].tolist()
    gekozen_locatie = st.selectbox("Kies locatie", locatie_options, index=0)

    ldf = tmp[tmp["Locatie"] == gekozen_locatie].copy()

    # Trend per maand
    ldf["Maand"] = ldf["Datum"].apply(_to_month)
    per_maand = (
        ldf[ldf["Maand"] != ""]
        .groupby("Maand")
        .size()
        .reset_index(name="Aantal")
        .sort_values("Maand")
    )

    cL, cR = st.columns([1.2, 1.0], gap="large")

    with cL:
        st.markdown("#### Schade per maand")
        if per_maand.empty:
            st.caption("Geen geldige datums om per maand te groeperen.")
        else:
            st.dataframe(per_maand, use_container_width=True, hide_index=True)
            st.bar_chart(per_maand.set_index("Maand")["Aantal"])

    with cR:
        st.markdown("#### Breakdown (top 10)")

        per_type = (
            ldf.groupby("type")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.caption("Type")
        st.dataframe(per_type, use_container_width=True, hide_index=True)

        per_voertuig = (
            ldf.groupby("voertuig")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.caption("Voertuig")
        st.dataframe(per_voertuig, use_container_width=True, hide_index=True)

        per_teamcoach = (
            ldf.groupby("teamcoach")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.caption("Teamcoach")
        st.dataframe(per_teamcoach, use_container_width=True, hide_index=True)

    st.markdown("#### Detail-lijst (laatste 200)")

    detail_cols = [
        c for c in
        ["Datum", "type", "voertuig", "bus/tram", "teamcoach", "volledige naam", "personeelsnr", "Link"]
        if c in ldf.columns
    ]
    details = ldf[detail_cols].copy()

    if "Datum" in details.columns:
        details["Datum"] = details["Datum"].apply(format_ddmmyyyy)

    # Sorteer op datum (nieuwste eerst) indien mogelijk
    try:
        sort_ts = pd.to_datetime(ldf["Datum"].astype(str), dayfirst=True, errors="coerce")
        details["_sort"] = sort_ts
        details = details.sort_values("_sort", ascending=False).drop(columns=["_sort"])
    except Exception:
        pass

    if "Link" in details.columns:
        details["Link"] = details["Link"].replace({"": None})

    column_config = {}
    if "Link" in details.columns:
        column_config["Link"] = st.column_config.LinkColumn("Open EAF", display_text="Open EAF", width="small")

    st.dataframe(
        details.head(200),
        use_container_width=True,
        hide_index=True,
        column_config=column_config if column_config else None,
    )

elif current_page == "analyse":
    st.subheader("Analyse")

    if df_schade_view.empty:
        st.info("Geen schadegegevens beschikbaar voor deze selectie.")
        st.stop()

    # ----------------------------
    # Helpers
    # ----------------------------
    def to_dt(v):
        return pd.to_datetime(v, dayfirst=True, errors="coerce")

    tmp = df_schade_view.copy()
    tmp["_dt"] = tmp["Datum"].apply(to_dt)
    tmp = tmp.dropna(subset=["_dt"])

    if tmp.empty:
        st.warning("Geen geldige datums gevonden om analyse te maken.")
        st.stop()

    # Normaliseer kernkolommen
    for c in ["type", "Locatie", "voertuig", "teamcoach"]:
        if c in tmp.columns:
            tmp[c] = (
                tmp[c]
                .fillna("")
                .astype(str)
                .str.strip()
                .replace("", "(onbekend)")
            )

    # ----------------------------
    # 1) Evolutie doorheen de tijd
    # ----------------------------
    st.markdown("## 📈 Evolutie doorheen de tijd")

    granularity = st.selectbox("Groeperen per", ["Maand", "Kwartaal"], index=0)

    if granularity == "Maand":
        tmp["Periode"] = tmp["_dt"].dt.to_period("M").astype(str)
    else:
        tmp["Periode"] = tmp["_dt"].dt.to_period("Q").astype(str)

    evolutie = (
        tmp.groupby("Periode")
        .size()
        .reset_index(name="Aantal schadegevallen")
        .sort_values("Periode")
    )

    c1, c2 = st.columns([1.1, 1.0])
    with c1:
        st.dataframe(evolutie, use_container_width=True, hide_index=True)
    with c2:
        st.bar_chart(evolutie.set_index("Periode")["Aantal schadegevallen"])

    st.divider()

    # ----------------------------
    # 2) Verdeling per type
    # ----------------------------
    st.markdown("## 🧩 Verdeling per type")

    per_type = (
        tmp.groupby("type")
        .size()
        .reset_index(name="Aantal")
        .sort_values("Aantal", ascending=False)
    )
    total = per_type["Aantal"].sum()
    per_type["Aandeel (%)"] = (per_type["Aantal"] / total * 100).round(1)

    st.dataframe(
        per_type.head(10),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Aandeel (%)": st.column_config.NumberColumn("Aandeel (%)", format="%.1f"),
        },
    )

    st.divider()

    # ----------------------------
    # 3) Hotspot-combinaties
    # ----------------------------
    st.markdown("## 🔥 Hotspot-combinaties")

    c1, c2 = st.columns(2)

    with c1:
        st.markdown("### Locatie × Type")
        loc_type = (
            tmp.groupby(["Locatie", "type"])
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.dataframe(loc_type, use_container_width=True, hide_index=True)

    with c2:
        st.markdown("### Voertuig × Type")
        veh_type = (
            tmp.groupby(["voertuig", "type"])
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.dataframe(veh_type, use_container_width=True, hide_index=True)

    st.divider()

    # ----------------------------
    # 4) Recente signalen – laatste 6 maanden
    # ----------------------------
    st.markdown("## 🚨 Recente signalen (laatste 6 maanden)")

    max_dt = tmp["_dt"].max()
    cutoff = max_dt - pd.DateOffset(months=6)

    recent = tmp[tmp["_dt"] >= cutoff].copy()

    if recent.empty:
        st.caption("Geen schadegevallen in de laatste 6 maanden.")
        st.stop()

    c1, c2, c3 = st.columns(3)

    with c1:
        st.markdown("### Locaties")
        recent_loc = (
            recent.groupby("Locatie")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.dataframe(recent_loc, use_container_width=True, hide_index=True)

    with c2:
        st.markdown("### Voertuigen")
        recent_veh = (
            recent.groupby("voertuig")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.dataframe(recent_veh, use_container_width=True, hide_index=True)

    with c3:
        st.markdown("### Types")
        recent_type = (
            recent.groupby("type")
            .size()
            .reset_index(name="Aantal")
            .sort_values("Aantal", ascending=False)
            .head(10)
        )
        st.dataframe(recent_type, use_container_width=True, hide_index=True)

    st.caption(
        f"Analyseperiode: {cutoff.strftime('%d-%m-%Y')} → {max_dt.strftime('%d-%m-%Y')}"
    )


            
elif current_page == "coaching":
    st.subheader("Coaching")

    # ----------------------------
    # Basis data (al geladen bovenaan)
    # df_coach_tab : geplande coaching (geen jaarfilter)
    # df_coach_voltooid_view : voltooide coaching (wel jaarfilter via _jaar)
    # df_schade_view : schade (wel jaarfilter via topbar)
    # ----------------------------

    # ----------------------------
    # Helpers
    # ----------------------------
    def _to_dt(v):
        return pd.to_datetime(v, dayfirst=True, errors="coerce")

    def _mode_or_empty(s: pd.Series) -> str:
        s = s.dropna().astype(str).str.strip()
        s = s[s != ""]
        if s.empty:
            return ""
        return s.value_counts().index[0]

    # ----------------------------
    # Normaliseer coaching tab
    # ----------------------------
    planned = df_coach_tab.copy()
    if planned.empty:
        planned = pd.DataFrame(columns=["nummer", "Chauffeurnaam", "Info", "_search"])

    planned["nummer"] = planned.get("nummer", "").fillna("").astype(str).apply(clean_id)
    planned["Chauffeurnaam"] = planned.get("Chauffeurnaam", "").fillna("").astype(str).str.strip()
    planned["Info"] = planned.get("Info", "").fillna("").astype(str).str.strip()

    planned["nummer"] = planned["nummer"].replace("", "(onbekend)")
    planned["Chauffeurnaam"] = planned["Chauffeurnaam"].replace("", "(onbekend)")
    planned["Info"] = planned["Info"].replace("", "")

    # ----------------------------
    # Bouw schade-statistieken op gekozen jaarfilter (df_schade_view)
    # ----------------------------
    schade = df_schade_view.copy()

    if schade.empty:
        schade_stats = pd.DataFrame(columns=["personeelsnr", "Schade (jaar)", "Laatste schade datum", "Top locatie", "Top type"])
    else:
        schade["personeelsnr"] = schade["personeelsnr"].apply(clean_id)
        schade["_dt"] = schade["Datum"].apply(_to_dt)

        schade["Locatie"] = schade["Locatie"].fillna("").astype(str).str.strip().replace("", "(onbekend)")
        schade["type"] = schade["type"].fillna("").astype(str).str.strip().replace("", "(onbekend)")

        schade_stats = (
            schade.groupby("personeelsnr", dropna=False)
            .agg(
                **{
                    "Schade (jaar)": ("personeelsnr", "size"),
                    "Laatste schade datum": ("_dt", "max"),
                    "Top locatie": ("Locatie", _mode_or_empty),
                    "Top type": ("type", _mode_or_empty),
                }
            )
            .reset_index()
        )

        schade_stats["Laatste schade datum"] = pd.to_datetime(schade_stats["Laatste schade datum"], errors="coerce")
        schade_stats["Laatste schade datum"] = schade_stats["Laatste schade datum"].dt.strftime("%d-%m-%Y").fillna("")

    # ----------------------------
    # Merge: planned coaching + schade stats (jaarfilter)
    # ----------------------------
    queue = planned.copy()
    # In df_coach_tab heet id "nummer", in schade "personeelsnr"
    queue = queue.merge(
        schade_stats,
        left_on="nummer",
        right_on="personeelsnr",
        how="left"
    )

    if "personeelsnr" in queue.columns:
        queue = queue.drop(columns=["personeelsnr"])

    # Vul NaNs
    if "Schade (jaar)" in queue.columns:
        queue["Schade (jaar)"] = queue["Schade (jaar)"].fillna(0).astype(int)
    else:
        queue["Schade (jaar)"] = 0

    for c in ["Laatste schade datum", "Top locatie", "Top type"]:
        if c not in queue.columns:
            queue[c] = ""
        queue[c] = queue[c].fillna("").astype(str)

    # ----------------------------
    # KPI's
    # ----------------------------
    totaal_gepland = len(queue)
    uniek_gepland = queue["nummer"].nunique() if "nummer" in queue.columns else 0
    totaal_voltooid = len(df_coach_voltooid_view) if df_coach_voltooid_view is not None else 0
    totaal_schade = len(df_schade_view)

    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Geplande coachings (rijen)", f"{totaal_gepland}")
    k2.metric("Geplande coachings (unieke P-nr)", f"{uniek_gepland}")
    k3.metric("Voltooide coachings (in selectie)", f"{totaal_voltooid}")
    k4.metric("Schadegevallen (in selectie)", f"{totaal_schade}")

    st.divider()

    # ----------------------------
    # Filters + sortering
    # ----------------------------
    c1, c2, c3 = st.columns([1.4, 1.0, 1.0])
    with c1:
        q = st.text_input("Zoek (P-nr / naam / info)", placeholder="Typ om te filteren…").strip().lower()
    with c2:
        sort_opt = st.selectbox(
            "Sorteren op",
            ["Schade (jaar) ↓", "Schade (jaar) ↑", "Naam A→Z", "Naam Z→A"],
            index=0
        )
    with c3:
        min_schade = st.slider("Minimum schade (jaar)", 0, 50, 0)

    filtered = queue.copy()

    if q:
        # veilige search over nummer/naam/info
        s = (
            filtered["nummer"].fillna("").astype(str) + " " +
            filtered["Chauffeurnaam"].fillna("").astype(str) + " " +
            filtered["Info"].fillna("").astype(str)
        ).str.lower()
        filtered = filtered[s.str.contains(re.escape(q), na=False)].copy()

    filtered = filtered[filtered["Schade (jaar)"] >= min_schade].copy()

    if sort_opt == "Schade (jaar) ↓":
        filtered = filtered.sort_values(["Schade (jaar)", "Chauffeurnaam"], ascending=[False, True])
    elif sort_opt == "Schade (jaar) ↑":
        filtered = filtered.sort_values(["Schade (jaar)", "Chauffeurnaam"], ascending=[True, True])
    elif sort_opt == "Naam A→Z":
        filtered = filtered.sort_values(["Chauffeurnaam", "Schade (jaar)"], ascending=[True, False])
    else:
        filtered = filtered.sort_values(["Chauffeurnaam", "Schade (jaar)"], ascending=[False, False])

    # ----------------------------
    # Werkqueue + dossier
    # ----------------------------
    left, right = st.columns([1.25, 1.0], gap="large")

    with left:
        st.markdown("### 📋 Werkqueue: Geplande coachings")

        show_cols = ["nummer", "Chauffeurnaam", "Info", "Schade (jaar)", "Laatste schade datum", "Top locatie", "Top type"]
        show_cols = [c for c in show_cols if c in filtered.columns]

        st.dataframe(
            filtered[show_cols].head(500),
            use_container_width=True,
            hide_index=True,
            column_config={
                "nummer": st.column_config.TextColumn("P-nr", width="small"),
                "Chauffeurnaam": st.column_config.TextColumn("Naam", width="medium"),
                "Info": st.column_config.TextColumn("Opmerking", width="large"),
                "Schade (jaar)": st.column_config.NumberColumn("Schade (jaar)", width="small"),
                "Laatste schade datum": st.column_config.TextColumn("Laatste schade", width="small"),
                "Top locatie": st.column_config.TextColumn("Top locatie", width="medium"),
                "Top type": st.column_config.TextColumn("Top type", width="medium"),
            },
        )

        # Export
        csv_bytes = filtered[show_cols].to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇️ Download werkqueue (CSV)",
            data=csv_bytes,
            file_name="werkqueue_coaching.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with right:
        st.markdown("### 🧾 Dossier: gekozen chauffeur")

        if filtered.empty:
            st.caption("Geen resultaten binnen deze filters.")
            st.stop()

        # Selecteer chauffeur uit gefilterde queue
        options = (
            filtered[["nummer", "Chauffeurnaam"]]
            .fillna("")
            .astype(str)
            .agg(" — ".join, axis=1)
            .tolist()
        )
        chosen = st.selectbox("Kies chauffeur", options, index=0)

        chosen_nummer = chosen.split(" — ")[0].strip()
        chosen_name = chosen.split(" — ")[1].strip() if " — " in chosen else ""

        # Geplande coaching info (alle rijen voor dit nummer)
        planned_person = queue[queue["nummer"] == chosen_nummer].copy()

        st.markdown("#### Geplande coaching")
        if planned_person.empty:
            st.caption("Geen geplande coaching gevonden voor deze chauffeur.")
        else:
            # Toon alle opmerkingen (soms meerdere rijen)
            for i, row in enumerate(planned_person.head(20).itertuples(index=False), start=1):
                info = getattr(row, "Info", "")
                st.write(f"**{i}.** {info}" if info else f"**{i}.** (geen opmerking)")

            # Schade-samenvatting
            st.markdown("#### Schade (in gekozen jaarfilter)")
            schade_row = planned_person.iloc[0]
            st.write(f"**Schade (jaar):** {int(schade_row.get('Schade (jaar)', 0))}")
            st.write(f"**Laatste schade datum:** {schade_row.get('Laatste schade datum', '')}")
            st.write(f"**Top locatie:** {schade_row.get('Top locatie', '')}")
            st.write(f"**Top type:** {schade_row.get('Top type', '')}")

        # Voltooide coachings (jaarfilter via df_coach_voltooid_view)
        st.markdown("#### Voltooide coachings (in selectie)")
        done = df_coach_voltooid_view.copy()
        if done is None or done.empty:
            st.caption("Geen voltooide coachings beschikbaar.")
        else:
            done["nummer"] = done["nummer"].apply(clean_id)
            done_person = done[done["nummer"] == chosen_nummer].copy()

            if done_person.empty:
                st.caption("Geen voltooide coachings gevonden voor deze chauffeur (binnen selectie).")
            else:
                # Format datum
                if "Datum" in done_person.columns:
                    done_person["Datum"] = done_person["Datum"].apply(format_ddmmyyyy)

                display = done_person[["Datum", "Info"]].copy() if "Datum" in done_person.columns else done_person[["Info"]].copy()
                render_html_table(
                    display.head(50),
                    col_order=list(display.columns),
                    col_widths={"Datum": "120px", "Info": "auto"},
                    max_height_px=360,
                )

        # Detail schadegevallen voor gekozen chauffeur (jaarfilter)
        st.markdown("#### Detail schadegevallen (laatste 200)")
        if df_schade_view.empty:
            st.caption("Geen schadegegevens in deze selectie.")
        else:
            sdf = df_schade_view.copy()
            sdf["personeelsnr"] = sdf["personeelsnr"].apply(clean_id)
            sdf = sdf[sdf["personeelsnr"] == chosen_nummer].copy()

            if sdf.empty:
                st.caption("Geen schadegevallen gevonden voor deze chauffeur (binnen selectie).")
            else:
                detail_cols = [c for c in ["Datum", "Locatie", "type", "voertuig", "bus/tram", "teamcoach", "Link"] if c in sdf.columns]
                details = sdf[detail_cols].copy()

                if "Datum" in details.columns:
                    details["Datum"] = details["Datum"].apply(format_ddmmyyyy)

                # Sorteer op datum (nieuwste eerst) indien mogelijk
                try:
                    sort_ts = pd.to_datetime(sdf["Datum"].astype(str), dayfirst=True, errors="coerce")
                    details["_sort"] = sort_ts
                    details = details.sort_values("_sort", ascending=False).drop(columns=["_sort"])
                except Exception:
                    pass

                if "Link" in details.columns:
                    details["Link"] = details["Link"].replace({"": None})

                column_config = {}
                if "Link" in details.columns:
                    column_config["Link"] = st.column_config.LinkColumn("Open EAF", display_text="Open EAF", width="small")

                st.dataframe(
                    details.head(200),
                    use_container_width=True,
                    hide_index=True,
                    column_config=column_config if column_config else None,
                )

    st.divider()

    # ----------------------------
    # Overzicht voltooide coachings (onderaan)
    # ----------------------------
    st.markdown("### ✅ Overzicht voltooide coachings (in selectie)")

    if df_coach_voltooid_view is None or df_coach_voltooid_view.empty:
        st.caption("Geen voltooide coachings beschikbaar voor deze selectie.")
    else:
        done_all = df_coach_voltooid_view.copy()
        done_all["nummer"] = done_all["nummer"].apply(clean_id)
        if "Datum" in done_all.columns:
            done_all["Datum"] = done_all["Datum"].apply(format_ddmmyyyy)

        q2 = st.text_input("Zoek in voltooide coachings (P-nr / naam / info)", placeholder="Typ om te filteren…", key="done_search").strip().lower()

        if q2:
            s2 = (
                done_all["nummer"].fillna("").astype(str) + " " +
                done_all.get("Chauffeurnaam", "").fillna("").astype(str) + " " +
                done_all.get("Info", "").fillna("").astype(str)
            ).str.lower()
            done_all = done_all[s2.str.contains(re.escape(q2), na=False)].copy()

        show_done_cols = [c for c in ["nummer", "Chauffeurnaam", "Datum", "Info"] if c in done_all.columns]
        render_html_table(
            done_all[show_done_cols].head(300),
            col_order=show_done_cols,
            col_widths={"nummer": "90px", "Chauffeurnaam": "180px", "Datum": "120px", "Info": "auto"},
            max_height_px=520,
        )

        csv_done = done_all[show_done_cols].to_csv(index=False).encode("utf-8")
        st.download_button(
            "⬇️ Download voltooide coachings (CSV)",
            data=csv_done,
            file_name="voltooide_coachings.csv",
            mime="text/csv",
            use_container_width=True,
        )
