from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, JSONResponse
from pathlib import Path
from typing import Optional, List, Dict, Any
import datetime as dt
import re

import openpyxl

APP_DIR = Path(__file__).parent

XLSM_PATH = APP_DIR / "schade met macro.xlsm"
INDEX_HTML = APP_DIR / "index.html"
LOGO_PNG = APP_DIR / "logo.png"

SHEET_NAME = "BRON"
COLS = [
    "personeelsnr",
    "volledige naam",
    "Datum",
    "Link",
    "Locatie",
    "voertuig",
    "bus/tram",
    "type",
]

app = FastAPI(title="OT Gent API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # later strenger maken
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

_cache_rows: Optional[List[Dict[str, Any]]] = None
_cache_mtime: Optional[float] = None


def _parse_year(value) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.year

    s = str(value).strip()
    if not s:
        return None

    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if m:
        return int(m.group(3))

    m2 = re.match(r"^(\d{4})[/-](\d{1,2})[/-](\d{1,2})", s)
    if m2:
        return int(m2.group(1))

    try:
        d = dt.datetime.fromisoformat(s)
        return d.year
    except Exception:
        return None


def _norm(h: Any) -> str:
    return str(h).strip().lower()


def _load_bron_rows() -> List[Dict[str, Any]]:
    global _cache_rows, _cache_mtime

    if not XLSM_PATH.exists():
        raise FileNotFoundError(f"Bestand niet gevonden: {XLSM_PATH.name} (plaats in dezelfde map als server.py)")

    mtime = XLSM_PATH.stat().st_mtime
    if _cache_rows is not None and _cache_mtime == mtime:
        return _cache_rows

    wb = openpyxl.load_workbook(XLSM_PATH, data_only=True, keep_vba=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Tabblad '{SHEET_NAME}' niet gevonden in {XLSM_PATH.name}")

    ws = wb[SHEET_NAME]

    header = [c.value for c in ws[1]]
    header_map = {_norm(h): idx for idx, h in enumerate(header)}

    def find_idx(col: str) -> Optional[int]:
        key = col.strip().lower()
        if key in header_map:
            return header_map[key]
        if col == "bus/tram":
            for alt in ["bus/ tram", "bus / tram", "bus - tram"]:
                if alt in header_map:
                    return header_map[alt]
        if col == "volledige naam":
            for alt in ["naam", "volledige naam.", "volledige naam "]:
                if alt in header_map:
                    return header_map[alt]
        return None

    idx_map = {col: find_idx(col) for col in COLS}

    rows: List[Dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        obj: Dict[str, Any] = {}
        any_val = False

        for col in COLS:
            i = idx_map.get(col)
            val = r[i] if i is not None and i < len(r) else None

            if val is not None and str(val).strip() != "":
                any_val = True

            if col == "Datum" and isinstance(val, (dt.date, dt.datetime)):
                val = val.isoformat()

            obj[col] = val

        if any_val:
            obj["_jaar"] = _parse_year(obj.get("Datum"))
            pn = str(obj.get("personeelsnr") or "")
            nm = str(obj.get("volledige naam") or "")
            vh = str(obj.get("voertuig") or "")
            obj["_search"] = (pn + " " + nm + " " + vh).lower()
            rows.append(obj)

    _cache_rows = rows
    _cache_mtime = mtime
    return rows


@app.get("/")
def index():
    return FileResponse(INDEX_HTML)


@app.get("/logo.png")
def logo():
    return FileResponse(LOGO_PNG)


@app.get("/api/jaren")
def jaren():
    rows = _load_bron_rows()
    years = sorted({r.get("_jaar") for r in rows if r.get("_jaar")}, reverse=True)
    return {"years": years}


@app.get("/api/bron")
def bron(
    jaar: Optional[int] = Query(default=None),
    q: Optional[str] = Query(default=None),
    limit: int = Query(default=200, ge=1, le=5000),
):
    rows = _load_bron_rows()

    if jaar is not None:
        rows = [r for r in rows if r.get("_jaar") == jaar]

    if q:
        qq = q.strip().lower()
        if qq:
            rows = [r for r in rows if qq in (r.get("_search") or "")]

    out = [{k: r.get(k) for k in COLS} for r in rows[:limit]]
    return {"count": len(rows), "items": out}


@app.get("/api/health")
def health():
    try:
        rows = _load_bron_rows()
        return {"ok": True, "rows": len(rows), "file": XLSM_PATH.name, "sheet": SHEET_NAME}
    except Exception as e:
        return JSONResponse(status_code=500, content={"ok": False, "error": str(e)})
